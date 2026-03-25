import json
import logging
import os
import re
import zipfile
from typing import Any, Callable, Dict, List, Optional
from xml.etree import ElementTree as ET

from test_platform.core.data_generators.test_case_exporter import parse_test_cases_from_text
from test_platform.core.document_processing.document_reader import read_document
from test_platform.core.services.result_contracts import (
    DEFAULT_STEP_EXPECTED,
    build_case_suite,
    build_case_suite_markdown,
    build_test_case_review_payload,
)
from test_platform.utils.json_utils import parse_json_markdown


logger = logging.getLogger(__name__)


class TestCaseReviewService:
    """负责测试用例导入解析、统一归一化与评审结果构建。"""

    EXCEL_COLUMN_ALIASES = {
        "title": {"用例名称", "案例名称", "测试用例", "标题", "用例标题"},
        "module": {"所属模块", "模块", "功能模块", "模块名称"},
        "precondition": {"前置条件", "前置", "预置条件"},
        "steps": {"步骤描述", "步骤", "操作步骤"},
        "expected": {"预期结果", "预期", "结果"},
        "priority": {"用例等级", "优先级", "priority", "级别"},
        "tags": {"标签", "tag", "tags"},
        "remark": {"备注", "说明", "remark"},
    }

    def resolve_case_suite(
        self,
        params: Optional[Dict[str, Any]],
        uploaded_paths: List[str],
    ) -> Dict[str, Any]:
        normalized_params = params or {}

        case_suite = normalized_params.get("case_suite")
        if isinstance(case_suite, dict) and isinstance(case_suite.get("items"), list):
            return build_case_suite(
                case_suite.get("items") or [],
                summary=str(case_suite.get("summary") or "已载入测试用例。").strip() or "已载入测试用例。",
            )

        for key in ("case_result", "case_content"):
            raw_text = str(normalized_params.get(key) or "").strip()
            if raw_text:
                return self._build_suite_from_text(raw_text)

        file_indexes = self._resolve_case_file_indexes(normalized_params, uploaded_paths)
        file_paths = [
            uploaded_paths[index]
            for index in file_indexes
            if 0 <= index < len(uploaded_paths)
        ]
        if not file_paths:
            raise ValueError("请提供测试用例内容、测试用例结果，或上传测试用例文件")

        merged_cases: List[Dict[str, Any]] = []
        for file_path in file_paths:
            merged_cases.extend(self._parse_cases_from_file(file_path))

        if not merged_cases:
            raise ValueError("未识别到有效测试用例")

        return build_case_suite(
            merged_cases,
            summary=f"已载入 {len(merged_cases)} 条测试用例，等待评审。",
        )

    def review_cases(
        self,
        requirement_text: str,
        suite: Dict[str, Any],
        reviewer: Optional[Callable[[str], str]] = None,
        extra_prompt: str = "",
    ) -> Dict[str, Any]:
        prompt = self._build_review_prompt(requirement_text, suite, extra_prompt=extra_prompt)
        raw_response = reviewer(prompt) if reviewer else ""
        parsed = parse_json_markdown(raw_response) if raw_response else None

        if isinstance(parsed, dict):
            return build_test_case_review_payload(
                summary=str(parsed.get("summary") or "").strip() or self._build_summary_text(suite, parsed.get("findings")),
                findings=parsed.get("findings") or [],
                reviewed_cases=parsed.get("reviewed_cases") or self._build_deterministic_reviewed_cases(suite),
                revised_suite=parsed.get("revised_suite") or suite,
                markdown=str(parsed.get("markdown") or "").strip(),
            )

        fallback_findings = self._build_deterministic_findings(suite)
        fallback_reviewed_cases = self._build_deterministic_reviewed_cases(suite)
        return build_test_case_review_payload(
            summary=self._build_summary_text(suite, fallback_findings),
            findings=fallback_findings,
            reviewed_cases=fallback_reviewed_cases,
            revised_suite=suite,
        )

    def _resolve_case_file_indexes(
        self,
        params: Dict[str, Any],
        uploaded_paths: List[str],
    ) -> List[int]:
        raw_indexes = params.get("case_file_indexes")
        normalized_indexes = self._normalize_index_list(raw_indexes)
        if normalized_indexes:
            return normalized_indexes

        requirement_indexes = set(self._normalize_index_list(params.get("requirement_file_indexes")))
        if requirement_indexes:
            return [index for index in range(len(uploaded_paths)) if index not in requirement_indexes]

        return list(range(len(uploaded_paths)))

    def _normalize_index_list(self, value: Any) -> List[int]:
        if not isinstance(value, list):
            return []

        normalized: List[int] = []
        for item in value:
            try:
                normalized.append(int(item))
            except (TypeError, ValueError):
                continue
        return normalized

    def _build_suite_from_text(self, raw_text: str) -> Dict[str, Any]:
        parsed_payload = parse_json_markdown(raw_text)
        if isinstance(parsed_payload, dict) and isinstance(parsed_payload.get("items"), list):
            normalized_summary = str(parsed_payload.get("summary") or "").strip()
            return build_case_suite(
                parsed_payload.get("items") or [],
                summary=normalized_summary or None,
            )
        if isinstance(parsed_payload, list):
            return build_case_suite(parsed_payload)

        cases = parse_test_cases_from_text(raw_text)
        if not cases:
            raise ValueError("未识别到有效测试用例")
        return build_case_suite(
            cases,
            summary=f"已载入 {len(cases)} 条测试用例，等待评审。",
        )

    def _parse_cases_from_file(self, file_path: str) -> List[Dict[str, Any]]:
        ext = os.path.splitext(file_path)[1].lower()
        if ext in {".xlsx", ".xls"}:
            return self._parse_cases_from_excel(file_path)
        if ext == ".xmind":
            return self._parse_cases_from_xmind(file_path)

        text, is_pdf = read_document(file_path)
        if is_pdf:
            raise ValueError("测试用例导入暂不支持 PDF，请转换为 Markdown、Excel 或 XMind")

        cases = parse_test_cases_from_text(text)
        if cases:
            return cases
        raise ValueError(f"文件 {os.path.basename(file_path)} 中未识别到有效测试用例")

    def _parse_cases_from_excel(self, file_path: str) -> List[Dict[str, Any]]:
        from openpyxl import load_workbook

        workbook = load_workbook(file_path, read_only=True, data_only=True)
        parsed_cases: List[Dict[str, Any]] = []
        try:
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                rows = list(worksheet.iter_rows(values_only=True))
                if not rows:
                    continue

                header_row_index = self._locate_header_row(rows)
                if header_row_index is None:
                    continue

                header_map = self._build_excel_header_map(rows[header_row_index])
                if "title" not in header_map:
                    raise ValueError("Excel 列头无法识别，至少需要包含“用例名称”列")

                for row in rows[header_row_index + 1:]:
                    row_values = [self._to_excel_cell_text(cell) for cell in row]
                    if not any(row_values):
                        continue

                    title = self._get_excel_value(row_values, header_map, "title")
                    if not title:
                        continue

                    steps_text = self._get_excel_value(row_values, header_map, "steps")
                    expected_text = self._get_excel_value(row_values, header_map, "expected")
                    steps = self._build_steps_from_excel_text(steps_text, expected_text)
                    parsed_cases.append(
                        {
                            "module": self._get_excel_value(row_values, header_map, "module") or sheet_name,
                            "name": title,
                            "precondition": self._get_excel_value(row_values, header_map, "precondition"),
                            "priority": self._get_excel_value(row_values, header_map, "priority") or "P1",
                            "tags": self._get_excel_value(row_values, header_map, "tags"),
                            "remark": self._get_excel_value(row_values, header_map, "remark"),
                            "steps": steps,
                        }
                    )
        finally:
            workbook.close()

        if not parsed_cases:
            raise ValueError("未识别到有效测试用例")
        return parsed_cases

    def _locate_header_row(self, rows: List[Any]) -> Optional[int]:
        for index, row in enumerate(rows[:10]):
            header_map = self._build_excel_header_map(row)
            if "title" in header_map:
                return index
        return None

    def _build_excel_header_map(self, row: Any) -> Dict[str, int]:
        header_map: Dict[str, int] = {}
        for index, cell in enumerate(row or []):
            normalized = self._normalize_excel_header(cell)
            if not normalized:
                continue
            for field_name, aliases in self.EXCEL_COLUMN_ALIASES.items():
                if normalized in aliases:
                    header_map[field_name] = index
                    break
        return header_map

    def _normalize_excel_header(self, value: Any) -> str:
        return str(value or "").strip().replace("：", "").replace(" ", "")

    def _to_excel_cell_text(self, value: Any) -> str:
        return str(value or "").strip()

    def _get_excel_value(self, row_values: List[str], header_map: Dict[str, int], field_name: str) -> str:
        index = header_map.get(field_name)
        if index is None or index >= len(row_values):
            return ""
        return row_values[index].strip()

    def _build_steps_from_excel_text(self, steps_text: str, expected_text: str) -> List[Dict[str, str]]:
        step_lines = self._split_ordered_lines(steps_text) or ["待补充步骤"]
        expected_lines = self._split_ordered_lines(expected_text)

        if not expected_lines:
            expected_lines = [DEFAULT_STEP_EXPECTED] * len(step_lines)
        elif len(expected_lines) == 1 and len(step_lines) > 1:
            expected_lines = expected_lines * len(step_lines)
        elif len(expected_lines) < len(step_lines):
            expected_lines.extend([expected_lines[-1]] * (len(step_lines) - len(expected_lines)))

        return [
            {
                "step": step_lines[index],
                "expected": expected_lines[index] if index < len(expected_lines) else DEFAULT_STEP_EXPECTED,
            }
            for index in range(len(step_lines))
        ]

    def _split_ordered_lines(self, raw_text: str) -> List[str]:
        text = str(raw_text or "").strip()
        if not text:
            return []

        lines: List[str] = []
        for raw_line in re.split(r"[\r\n]+", text):
            line = raw_line.strip()
            if not line:
                continue
            line = re.sub(r"^\[\d+\]\s*", "", line)
            line = re.sub(r"^\d+\s*[.、]\s*", "", line)
            line = re.sub(r"^[\-*•]\s*", "", line)
            line = line.strip()
            if line:
                lines.append(line)
        return lines

    def _parse_cases_from_xmind(self, file_path: str) -> List[Dict[str, Any]]:
        with zipfile.ZipFile(file_path, "r") as archive:
            try:
                content = archive.read("content.xml")
            except KeyError as error:
                raise ValueError("XMind 文件缺少 content.xml，无法解析") from error

        namespace = {"x": "urn:xmind:xmap:xmlns:content:2.0"}
        root = ET.fromstring(content)
        root_topic = root.find(".//x:sheet/x:topic", namespace)
        if root_topic is None:
            raise ValueError("XMind 文件缺少根节点，无法解析")

        cases: List[Dict[str, Any]] = []
        for topic in self._get_attached_topics(root_topic, namespace):
            self._walk_xmind_topic(topic, namespace, parent_modules=[], cases=cases)

        if not cases:
            raise ValueError("XMind 节点结构不符合支持格式")
        return cases

    def _walk_xmind_topic(
        self,
        topic: ET.Element,
        namespace: Dict[str, str],
        parent_modules: List[str],
        cases: List[Dict[str, Any]],
    ) -> None:
        title = self._get_topic_title(topic, namespace)
        if not title:
            return

        if title.startswith("case：") or title.startswith("case:") or title.startswith("用例：") or title.startswith("案例："):
            cases.append(self._parse_xmind_case_topic(topic, namespace, parent_modules))
            return

        next_modules = parent_modules + [title]
        for child_topic in self._get_attached_topics(topic, namespace):
            self._walk_xmind_topic(child_topic, namespace, next_modules, cases)

    def _parse_xmind_case_topic(
        self,
        topic: ET.Element,
        namespace: Dict[str, str],
        parent_modules: List[str],
    ) -> Dict[str, Any]:
        raw_title = self._get_topic_title(topic, namespace)
        title = re.sub(r"^(case|用例|案例)[：:]\s*", "", raw_title, flags=re.IGNORECASE).strip() or "未命名用例"
        case_payload: Dict[str, Any] = {
            "module": "/".join(parent_modules) if parent_modules else "未分类",
            "name": title,
            "precondition": "",
            "priority": "P1",
            "tags": "",
            "remark": "",
            "steps": [],
        }

        for child_topic in self._get_attached_topics(topic, namespace):
            child_title = self._get_topic_title(child_topic, namespace)
            if child_title.startswith("前置条件："):
                case_payload["precondition"] = child_title.split("：", 1)[1].strip()
                continue
            if child_title.startswith("用例等级："):
                case_payload["priority"] = child_title.split("：", 1)[1].strip() or "P1"
                continue
            if child_title.startswith("标签："):
                case_payload["tags"] = child_title.split("：", 1)[1].strip()
                continue
            if child_title.startswith("备注："):
                case_payload["remark"] = child_title.split("：", 1)[1].strip()
                continue
            if child_title == "步骤描述":
                case_payload["steps"] = self._parse_xmind_step_topics(child_topic, namespace)

        return case_payload

    def _parse_xmind_step_topics(self, steps_topic: ET.Element, namespace: Dict[str, str]) -> List[Dict[str, str]]:
        steps: List[Dict[str, str]] = []
        for step_topic in self._get_attached_topics(steps_topic, namespace):
            step_title = self._get_topic_title(step_topic, namespace)
            if not step_title.startswith("步骤："):
                continue
            action = step_title.split("：", 1)[1].strip() or "待补充步骤"
            expected = DEFAULT_STEP_EXPECTED
            for child_topic in self._get_attached_topics(step_topic, namespace):
                child_title = self._get_topic_title(child_topic, namespace)
                if child_title.startswith("预期结果："):
                    expected = child_title.split("：", 1)[1].strip() or DEFAULT_STEP_EXPECTED
                    break
            steps.append({"step": action, "expected": expected})

        if not steps:
            return [{"step": "待补充步骤", "expected": DEFAULT_STEP_EXPECTED}]
        return steps

    def _get_attached_topics(self, topic: ET.Element, namespace: Dict[str, str]) -> List[ET.Element]:
        return list(topic.findall("./x:children/x:topics[@type='attached']/x:topic", namespace))

    def _get_topic_title(self, topic: ET.Element, namespace: Dict[str, str]) -> str:
        title_node = topic.find("./x:title", namespace)
        return str(title_node.text or "").strip() if title_node is not None else ""

    def _build_review_prompt(self, requirement_text: str, suite: Dict[str, Any], extra_prompt: str = "") -> str:
        suite_markdown = build_case_suite_markdown(suite)
        extra_prompt_block = f"\n[用户附加要求]\n{extra_prompt}\n" if extra_prompt else ""
        return f"""你是一名资深测试专家，请严格基于需求评审当前测试用例。

请重点检查：
1. 是否脱离原始需求
2. 是否遗漏关键验收点
3. 是否引入需求外场景
4. 步骤是否清晰、可执行
5. 优先级是否合理

必须输出 JSON 对象，字段固定为：
- summary
- findings
- reviewed_cases
- revised_suite
- markdown

输出约束：
- findings 中的 risk_level 只允许 H/M/L
- reviewed_cases 中的 verdict 只允许 pass/warning/fail
- reviewed_cases 中的 consistency 只允许 aligned/partial/deviated
- revised_suite 必须保持为可直接落回测试用例工作台的结构，至少包含 items 与 summary
- 所有内容必须使用简体中文
- 禁止输出 JSON 之外的解释文字
{extra_prompt_block}
[原始需求]
{requirement_text}

[当前测试用例]
{suite_markdown}
"""

    def _build_summary_text(self, suite: Dict[str, Any], findings: Any) -> str:
        items = suite.get("items") if isinstance(suite, dict) else []
        case_count = len(items) if isinstance(items, list) else 0
        finding_count = len(findings) if isinstance(findings, list) else 0
        high_risk_count = len(
            [item for item in (findings or []) if isinstance(item, dict) and str(item.get("risk_level")).upper() == "H"]
        ) if isinstance(findings, list) else 0
        return f"已评审 {case_count} 条测试用例，识别 {finding_count} 项问题，其中高风险 {high_risk_count} 项。"

    def _build_deterministic_findings(self, suite: Dict[str, Any]) -> List[Dict[str, Any]]:
        findings: List[Dict[str, Any]] = []
        items = suite.get("items") if isinstance(suite, dict) else []
        if not isinstance(items, list):
            return findings

        for item in items:
            if not isinstance(item, dict):
                continue
            case_id = str(item.get("id") or "")
            case_title = str(item.get("title") or "未命名用例")
            for step in item.get("steps") or []:
                if not isinstance(step, dict):
                    continue
                expected = str(step.get("expected") or "").strip()
                if not expected or expected == DEFAULT_STEP_EXPECTED:
                    findings.append(
                        {
                            "risk_level": "M",
                            "category": "步骤不清晰",
                            "related_case_ids": [case_id] if case_id else [],
                            "related_requirement_points": [],
                            "description": f"测试用例“{case_title}”存在预期结果不完整的步骤。",
                            "suggestion": "补充每一步的明确预期结果，避免执行时口径不一致。",
                        }
                    )
                    break
        return findings

    def _build_deterministic_reviewed_cases(self, suite: Dict[str, Any]) -> List[Dict[str, Any]]:
        reviewed_cases: List[Dict[str, Any]] = []
        items = suite.get("items") if isinstance(suite, dict) else []
        if not isinstance(items, list):
            return reviewed_cases

        for item in items:
            if not isinstance(item, dict):
                continue
            issues: List[str] = []
            suggestions: List[str] = []
            for step in item.get("steps") or []:
                if not isinstance(step, dict):
                    continue
                expected = str(step.get("expected") or "").strip()
                if not expected or expected == DEFAULT_STEP_EXPECTED:
                    issues.append("存在步骤预期结果缺失或过于笼统")
                    suggestions.append("补充步骤级预期结果")
                    break

            reviewed_cases.append(
                {
                    "case_id": str(item.get("id") or ""),
                    "title": str(item.get("title") or "未命名用例"),
                    "module": str(item.get("module") or "未分类"),
                    "verdict": "warning" if issues else "pass",
                    "consistency": "partial" if issues else "aligned",
                    "issues": issues,
                    "suggestions": suggestions,
                }
            )

        return reviewed_cases
