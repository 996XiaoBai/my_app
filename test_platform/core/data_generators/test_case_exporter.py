"""
测试用例导出模块。
支持将结构化测试用例数据导出为 Excel（TAPD 导入格式）和 XMind 格式。
"""
import json
import os
import logging
import zipfile
import re
from typing import List, Dict, Optional
from xml.etree import ElementTree as ET
from datetime import datetime
from test_platform.core.services.result_contracts import to_export_cases

logger = logging.getLogger(__name__)


def _clean_markdown_text(text: str) -> str:
    """清理 Markdown 文本中的占位符尖标签（如 <未登录用户>）与不必要的首尾空格。"""
    if not text:
        return text
    # 移除诸如 <...> 的尖括号，只保留里面的内容，避免被解析为 XML 或干扰肉眼阅读
    cleaned = re.sub(r'<([^>]+)>', r'\1', text)
    # 移除可能的加粗星号 ** 
    cleaned = cleaned.replace("**", "")
    # 移除中文和英文双引号
    cleaned = cleaned.replace('"', '').replace('“', '').replace('”', '')
    
    # 将中文和英文的分号、及可能包含的前后空格，统一替换为换行符
    # 如果文本中存在分号，则将其按分号分割，并加上 ① ② ③ 等序号
    if ';' in cleaned or '；' in cleaned:
        parts = re.split(r'\s*[;；]\s*', cleaned)
        numbered_parts = []
        _idx = 1
        for p in parts:
            if p.strip():
                # 为了美观，使用带圈数字 1-20，如果超出 20 则用 (21)
                circle_num = chr(9311 + _idx) if 1 <= _idx <= 20 else f"({_idx})"
                numbered_parts.append(f"{circle_num} {p.strip()}")
                _idx += 1
        cleaned = '\n'.join(numbered_parts)

    return cleaned.strip()

def _parse_markdown_test_cases(text: str) -> List[Dict]:
    """支持从纯净的 Markdown 层级缩进中提取测试用例框架。"""
    cases = []
    current_module = ""
    current_case = None
    lines = text.split('\n')
    
    for line in lines:
        line_s = line.strip()
        if not line_s:
            continue
            
        if line_s.startswith("## 模块：") or line_s.startswith("## 功能模块："):
            # 兼容带有 “功能模块：” 或只写 “模块：”
            current_module = _clean_markdown_text(re.sub(r'^##\s*(功能)?模块：', '', line_s))
        elif line_s.startswith("### case：") or line_s.startswith("### 案例：") or line_s.startswith("### 用例："):
            if current_case:
                cases.append(current_case)
            case_name = re.sub(r'^###\s*(case|案例|用例)：', '', line_s, flags=re.IGNORECASE)
            current_case = {
                "module": current_module,
                "name": _clean_markdown_text(case_name),
                "precondition": "",
                "priority": "P1",
                "tags": "",
                "remark": "",
                "steps": []
            }
        elif current_case:
            if line_s.startswith("- 前置条件："):
                current_case["precondition"] = _clean_markdown_text(line_s.replace("- 前置条件：", ""))
            elif line_s.startswith("- 用例等级：") or line_s.startswith("- 案例等级："):
                current_case["priority"] = _clean_markdown_text(re.sub(r'-\s*(用例|案例)等级：', '', line_s))
            elif line_s.startswith("- 步骤描述："):
                pass
            elif re.match(r'^\d+\.\s*步骤：', line_s):
                step_text = re.sub(r'^\d+\.\s*步骤：', '', line_s)
                current_case["steps"].append({"step": _clean_markdown_text(step_text), "expected": ""})
            elif line_s.startswith("- 预期结果：") or line_s.startswith("- 结果："):
                exp_text = re.sub(r'-\s*(预期)?结果：', '', line_s)
                if current_case["steps"]:
                    current_case["steps"][-1]["expected"] = _clean_markdown_text(exp_text)
                else:
                    current_case["steps"].append({"step": "", "expected": _clean_markdown_text(exp_text)})
                    
    if current_case:
        cases.append(current_case)
        
    return cases

def parse_test_cases_from_text(text: str) -> List[Dict]:
    """
    从 AI 生成的文本中解析出测试用例列表。
    支持 JSON 与我们定制的严格 Markdown 树两种格式提取。
    """
    # 优先支持新版结构化对象：{"items": [...], "summary": "..."}
    try:
        raw_data = json.loads(text)
        if isinstance(raw_data, dict) and isinstance(raw_data.get("items"), list):
            cases = to_export_cases(raw_data["items"])
            if cases:
                logger.info(f"✅ 成功解析 Case Suite {len(cases)} 条测试用例")
                return cases
        elif isinstance(raw_data, list):
            cases = to_export_cases(raw_data)
            if cases:
                logger.info(f"✅ 成功解析 JSON {len(cases)} 条测试用例")
                return cases
    except Exception:
        pass

    # 尝试从 markdown 代码块中提取 JSON
    json_match = re.search(r'```(?:json)?\s*\n([\s\S]*?)\n```', text)
    if json_match:
        json_str = json_match.group(1)
        try:
            raw_data = json.loads(json_str)
            if isinstance(raw_data, dict) and isinstance(raw_data.get("items"), list):
                cases = to_export_cases(raw_data["items"])
                if cases:
                    logger.info(f"✅ 成功解析 Case Suite {len(cases)} 条测试用例")
                    return cases
            elif isinstance(raw_data, list):
                cases = to_export_cases(raw_data)
                if cases:
                    logger.info(f"✅ 成功解析 JSON {len(cases)} 条测试用例")
                    return cases
        except json.JSONDecodeError:
            pass
            
    # 尝试直接解析
    start = text.find('[')
    end = text.rfind(']')
    if start != -1 and end != -1:
        try:
            cases = to_export_cases(json.loads(text[start:end + 1]))
            if cases:
                logger.info(f"✅ 成功解析 JSON {len(cases)} 条测试用例")
                return cases
        except json.JSONDecodeError:
            pass

    # 若 JSON 解析皆不中，则启动定制化 Markdown 锚点解析！
    markdown_cases = _parse_markdown_test_cases(text)
    if markdown_cases:
        logger.info(f"✅ 成功通过 Markdown 树解析 {len(markdown_cases)} 条测试用例")
        return markdown_cases
    else:
        logger.warning("未能从文本中提取任何格式的测试用例")
        return []


def export_to_excel(cases: List[Dict], output_path: str) -> str:
    """
    将测试用例导出为 Excel 文件（TAPD 导入格式）。
    
    列：用例名称 | 所属模块 | 标签 | 前置条件 | 步骤描述 | 预期结果 | 编辑模式 | 备注 | 用例等级
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例"
    
    # 表头
    headers = ["用例名称", "所属模块", "标签", "前置条件", "步骤描述", "预期结果", "编辑模式", "备注", "用例等级"]
    
    # 表头样式
    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    
    # 数据行
    data_align = Alignment(vertical="top", wrap_text=True)
    
    for row_idx, case in enumerate(cases, 2):
        steps = case.get("steps", [])
        
        # 格式化步骤描述
        step_text = "\n".join(
            [f"[{i+1}]{s.get('step', '')}" for i, s in enumerate(steps)]
        ) if steps else ""
        
        # 格式化预期结果
        expected_text = "\n".join(
            [f"[{i+1}]{s.get('expected', '')}" for i, s in enumerate(steps)]
        ) if steps else ""
        
        row_data = [
            case.get("name", ""),           # 用例名称
            case.get("module", ""),          # 所属模块
            case.get("tags", ""),            # 标签
            case.get("precondition", ""),    # 前置条件
            step_text,                       # 步骤描述
            expected_text,                   # 预期结果
            "STEP",                          # 编辑模式
            case.get("remark", ""),          # 备注
            case.get("priority", "P1"),      # 用例等级
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.alignment = data_align
            cell.border = thin_border
    
    # 设置列宽
    col_widths = [30, 25, 15, 25, 40, 40, 10, 20, 10]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width
    
    wb.save(output_path)
    logger.info(f"✅ Excel 导出成功: {output_path}（{len(cases)} 条用例）")
    return output_path


def export_to_xmind(cases: List[Dict], output_path: str) -> str:
    """
    将测试用例导出为 XMind 文件。
    
    结构：功能用例 → 模块 → 用例名称 → [前置条件/步骤描述/用例等级]
                                              └→ 步骤 → 预期结果
    """
    import time
    
    timestamp = str(int(time.time() * 1000))
    
    # 构建 XMind XML
    ns = "urn:xmind:xmap:xmlns:content:2.0"
    ET.register_namespace('', ns)
    
    xmap = ET.Element(f'{{{ns}}}xmap-content', {
        'version': '2.0',
        'timestamp': timestamp
    })
    
    sheet = ET.SubElement(xmap, f'{{{ns}}}sheet', {
        'id': _gen_id(),
        'timestamp': timestamp
    })
    
    # 根节点：功能用例
    root_topic = ET.SubElement(sheet, f'{{{ns}}}topic', {
        'id': _gen_id(),
        'timestamp': timestamp
    })
    title = ET.SubElement(root_topic, f'{{{ns}}}title')
    title.text = "功能用例"
    
    module_tree = _build_module_tree(cases)

    root_children = ET.SubElement(root_topic, f'{{{ns}}}children')
    root_topics = ET.SubElement(root_children, f'{{{ns}}}topics', {'type': 'attached'})
    _append_module_topics(ns, timestamp, root_topics, module_tree)
    
    # 写入 XMind 文件（本质是 ZIP）
    xml_str = b'<?xml version="1.0" encoding="UTF-8"?>' + ET.tostring(xmap, encoding='utf-8')
    
    meta_xml = '<?xml version="1.0" encoding="UTF-8"?><meta xmlns="urn:xmind:xmap:xmlns:meta:2.0" version="2.0"/>'
    manifest_xml = '<?xml version="1.0" encoding="UTF-8"?><manifest xmlns="urn:xmind:xmap:xmlns:manifest:1.0"><file-entry full-path="content.xml" media-type="text/xml"/><file-entry full-path="META-INF/" media-type=""/><file-entry full-path="META-INF/manifest.xml" media-type="text/xml"/></manifest>'
    
    with zipfile.ZipFile(output_path, 'w', zipfile.ZIP_DEFLATED) as zf:
        zf.writestr('content.xml', xml_str)
        zf.writestr('meta.xml', meta_xml)
        zf.writestr('META-INF/manifest.xml', manifest_xml)
    
    logger.info(f"✅ XMind 导出成功: {output_path}（{len(cases)} 条用例）")
    return output_path


def _gen_id() -> str:
    """生成简单的唯一 ID。"""
    import uuid
    return uuid.uuid4().hex[:26]


def _create_topic(ns: str, timestamp: str, title_text: str) -> ET.Element:
    """创建一个 XMind topic 节点。"""
    topic = ET.Element(f'{{{ns}}}topic', {
        'id': _gen_id(),
        'timestamp': timestamp
    })
    title = ET.SubElement(topic, f'{{{ns}}}title')
    title.text = title_text
    return topic


def _build_module_tree(cases: List[Dict]) -> Dict[str, Dict]:
    tree: Dict[str, Dict] = {"children": {}, "cases": []}
    for case in cases:
        module_path = str(case.get("module") or "未分类")
        parts = [part for part in module_path.split("/") if part] or ["未分类"]

        current = tree
        for part in parts:
            children = current.setdefault("children", {})
            current = children.setdefault(part, {"children": {}, "cases": []})
        current.setdefault("cases", []).append(case)
    return tree


def _append_module_topics(ns: str, timestamp: str, parent_topics: ET.Element, tree: Dict[str, Dict]) -> None:
    for module_name, module_node in tree.get("children", {}).items():
        module_topic = _create_topic(ns, timestamp, module_name)
        parent_topics.append(module_topic)

        module_children = ET.SubElement(module_topic, f'{{{ns}}}children')
        module_topics = ET.SubElement(module_children, f'{{{ns}}}topics', {'type': 'attached'})

        _append_module_topics(ns, timestamp, module_topics, module_node)

        for case in module_node.get("cases", []):
            _append_case_topic(ns, timestamp, module_topics, case)


def _append_case_topic(ns: str, timestamp: str, parent_topics: ET.Element, case: Dict) -> None:
    case_topic = _create_topic(ns, timestamp, f"case：{case.get('name', '')}")
    parent_topics.append(case_topic)

    case_children = ET.SubElement(case_topic, f'{{{ns}}}children')
    case_topics = ET.SubElement(case_children, f'{{{ns}}}topics', {'type': 'attached'})

    precondition = str(case.get("precondition") or "").strip()
    if precondition:
        case_topics.append(_create_topic(ns, timestamp, f"前置条件：{precondition}"))

    steps_topic = _create_topic(ns, timestamp, "步骤描述")
    case_topics.append(steps_topic)
    _append_step_topics(ns, timestamp, steps_topic, case.get("steps", []))

    case_topics.append(_create_topic(ns, timestamp, f"用例等级：{case.get('priority', 'P1')}"))

    tags = str(case.get("tags") or "").strip()
    if tags:
        case_topics.append(_create_topic(ns, timestamp, f"标签：{tags}"))

    remark = str(case.get("remark") or "").strip()
    if remark:
        case_topics.append(_create_topic(ns, timestamp, f"备注：{remark}"))


def _append_step_topics(ns: str, timestamp: str, steps_topic: ET.Element, steps: List[Dict]) -> None:
    steps_children = ET.SubElement(steps_topic, f'{{{ns}}}children')
    steps_topics = ET.SubElement(steps_children, f'{{{ns}}}topics', {'type': 'attached'})

    if not steps:
        steps = [{"step": "待补充步骤", "expected": "待补充预期结果"}]

    for step in steps:
        action = str(step.get("step") or "").strip() or "待补充步骤"
        expected = str(step.get("expected") or "").strip() or "待补充预期结果"

        step_topic = _create_topic(ns, timestamp, f"步骤：{action}")
        steps_topics.append(step_topic)

        step_children = ET.SubElement(step_topic, f'{{{ns}}}children')
        step_subtopics = ET.SubElement(step_children, f'{{{ns}}}topics', {'type': 'attached'})
        step_subtopics.append(_create_topic(ns, timestamp, f"预期结果：{expected}"))
